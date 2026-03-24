# from ws_ltp import run_ltp_socket
import datetime
import time
import os
import pandas as pd
import numpy as np
import pdb
import requests
from datetime import date, timedelta
import json
from datetime import time as dt_time
from ta.trend import EMAIndicator, SMAIndicator, ADXIndicator
from ta.momentum import RSIIndicator
from openpyxl import Workbook, load_workbook

from mock_engine import MockEngine	
engine = MockEngine()
from tools import Tools
tools = Tools()
from sdtools3 import SDTools3
sdtools = SDTools3()

# sdtools.get_instrument_file()

def send_telegram(message):
	try:
		import requests

		TELEGRAM_BOT_TOKEN = ""
		TELEGRAM_CHAT_ID = ""
		url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
		data = {"chat_id": TELEGRAM_CHAT_ID, "text": message}
		requests.post(url, data=data)
	except Exception as e:
		print(f"❌ Failed to send Telegram alert: {e}")

message1 = "Advanced Stock Trading Bot Started"
send_telegram(message1)
print("Advanced Stock Trading Bot Started")
'''
tools.get_instrument_file()

main_list = ["AARTIIND", "ABB", "ABCAPITAL", "ABFRL", "ADANIENSOL", "ADANIENT", "ADANIGREEN", "ADANIPORTS", "AMBUJACEM", "ANGELONE", "APLAPOLLO", "APOLLOHOSP", "ASHOKLEY", "ASIANPAINT", "ASTRAL", "ATGL", "AUBANK", "AUROPHARMA", "AXISBANK", "BAJAJ-AUTO", "BAJAJFINSV", "BAJFINANCE", "BALKRISIND", "BANDHANBNK", "BANKBARODA", "BEL", "BHARATFORG", "BHARTIARTL", "BIOCON", "BPCL", "BRITANNIA", "BSE", "BSOFT", "CAMS", "CDSL", "CGPOWER", "CHAMBLFERT", "CHOLAFIN", "CIPLA", "COALINDIA", "COFORGE", "CONCOR", "CROMPTON", "CUMMINSIND", "DABUR", "DALBHARAT", "DIVISLAB", "DIXON", "DLF", "DMART", "DRREDDY", "EICHERMOT", "ETERNAL", "EXIDEIND", "GAIL", "GLENMARK", "GODREJCP", "GODREJPROP", "GRANULES", "GRASIM", "HAL", "HAVELLS", "HCLTECH", "HDFCAMC", "HDFCBANK", "HDFCLIFE", "HEROMOTOCO", "HINDALCO", "HINDCOPPER", "HINDPETRO", "HINDUNILVR", "HINDZINC", "ICICIBANK", "ICICIGI", "ICICIPRULI", "IGL", "INDHOTEL", "INDIGO", "INDUSINDBK", "INDUSTOWER", "INFY", "IOC", "IRCTC", "IREDA", "ITC", "JINDALSTEL", "JIOFIN", "JSWSTEEL", "JUBLFOOD", "KAYNES", "KFINTECH", "KOTAKBANK", "KPITTECH", "LAURUSLABS", "LICHSGFIN", "LICI", "LODHA", "LT", "LTF", "LTIM", "LUPIN", "M&M", "M&MFIN", "MANAPPURAM", "MARICO", "MARUTI", "MAZDOCK", "MFSL", "MGL", "MOTHERSON", "MPHASIS", "MUTHOOTFIN", "NATIONALUM", "NAUKRI", "NESTLEIND", "NTPC", "NYKAA", "OBEROIRLTY", "OFSS", "OIL", "ONGC", "PEL", "PERSISTENT", "PETRONET", "PFC", "PIDILITIND", "PIIND", "PNBHOUSING", "POLYCAB", "POWERGRID", "RECLTD", "RELIANCE", "RVNL", "SBICARD", "SBILIFE", "SBIN", "SHRIRAMFIN", "SIEMENS", "SOLARINDS", "SRF", "SUNPHARMA", "TATACHEM", "TATACOMM", "TATACONSUM", "TATAELXSI", "TATAMOTORS", "TATAPOWER", "TATATECH", "TCS", "TECHM", "TIINDIA", "TITAN", "TORNTPHARM", "TORNTPOWER", "TRENT", "TVSMOTOR", "UNITDSPR", "UNOMINDA", "UPL", "VBL", "VEDL", "VOLTAS", "WIPRO", "ZYDUSLIFE"]

FnO_list = ["ABB", "ADANIENSOL", "ADANIENT", "ADANIGREEN", "ADANIPORTS", "AMBUJACEM", "APOLLOHOSP", "ASIANPAINT", "AXISBANK", "BAJAJFINSV", "BAJFINANCE", "BHARTIARTL", "BRITANNIA", "CHOLAFIN", "CIPLA", "COALINDIA", "DABUR", "DIVISLAB", "DLF", "DMART", "DRREDDY", "EICHERMOT", "GODREJCP", "GRASIM", "HAL", "HAVELLS", "HCLTECH", "HDFCBANK", "HDFCLIFE", "HEROMOTOCO", "HINDALCO", "HINDUNILVR", "ICICIBANK", "ICICIGI", "ICICIPRULI", "INDIGO", "INDUSINDBK", "INFY", "IRCTC", "JINDALSTEL", "JSWENERGY", "JSWSTEEL", "KOTAKBANK", "LICI", "LODHA", "LT", "LTIM", "M&M", "NAUKRI", "NESTLEIND", "NTPC", "PFC", "PIDILITIND", "RECLTD", "RELIANCE", "SBILIFE", "SBIN", "SHRIRAMFIN", "SIEMENS", "SUNPHARMA", "TATACONSUM", "TATAMOTORS", "TATAPOWER", "TCS", "TECHM", "TITAN", "TORNTPHARM", "TRENT", "TVSMOTOR", "UNITDSPR", "VBL", "VEDL", "ZYDUSLIFE"]

fo_list = ["INFY", "HINDALCO","NAUKRI"]

filename = "backtest_results.xlsx"

def reset_backtest_file(filename):
	try:
		# Try to load existing workbook
		book = load_workbook(filename)
		for sheet in book.sheetnames:
			std = book[sheet]
			book.remove(std)
	except FileNotFoundError:
		# If file does not exist, create new workbook
		book = Workbook()
		# Remove the default sheet
		book.remove(book.active)

	# Add fresh empty sheets
	bullish_engulfing = book.create_sheet("bullish_engulfing")
	bearish_engulfing = book.create_sheet("bearish_engulfing")
	hammer = book.create_sheet("hammer")
	shooting_star = book.create_sheet("shooting_star")

	headers = [
		"Stock Name", "Security ID", "Timestamp", "Open", "High", "Low", "Close",
		"Volume", "Datetime", "Entry"
	]
	bullish_engulfing.append(headers)
	bearish_engulfing.append(headers)
	hammer.append(headers)
	shooting_star.append(headers)

	book.save(filename)
	print(f"✅ Reset done: {filename}")
	
def append_df_to_excel(filename, df, sheet_name):
	from openpyxl import load_workbook

	if not os.path.isfile(filename):
		# Create new file and write the sheet
		with pd.ExcelWriter(filename, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name=sheet_name, index=False)
	else:
		# Open existing workbook
		book = load_workbook(filename)
		# Find the starting row for appending
		startrow = book[sheet_name].max_row if sheet_name in book.sheetnames else 0

		with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
			df.to_excel(writer, sheet_name=sheet_name, index=False, header=(startrow == 0), startrow=startrow)

current_time = datetime.datetime.now().time()
headers = ["Stock Name", "Security ID", "Timestamp", "Open", "High", "Low", "Close", "Volume", "Datetime", "Entry"]


while True:
	back_date = datetime.date(2025, 9, 3)
	open_high_dict, open_low_dict = sdtools.sd_open_high_low_dicts(stock_list=FnO_list, todays_date=back_date)
	# excel_path = "backtest_results.xlsx"  # your file name
	reset_backtest_file(filename)
	try:
		# --- CALL ENTRIES ---
		for stock_name, stock_data in open_low_dict.items():
			stock_id = stock_data["security_id"]
			df = sdtools.back_data(security_id=stock_id, interval=5, todays_date=back_date)
			call_df = sdtools.call_backtest_entries(df)
			call_df['stock_name'] = stock_name
			call_df['stock_id'] = stock_id
			cols = ['stock_name', 'stock_id'] + [c for c in call_df.columns if c not in ['stock_name', 'stock_id']]
			call_df = call_df[cols]
			call_df['datetime'] = pd.to_datetime(call_df['datetime']).dt.tz_localize(None)
			
			if not call_df.empty:
				call_df.columns = headers
				append_df_to_excel(filename, call_df, 'CALL_Entries')

			print("Call Side done")
			# print(entry_df), "ICICIGI", "ABB", "GRASIM", "TATAELXSI", "OFSS", "OFSS", "TECHM", "JUBLFOOD"
	except Exception as e:
		print(f"❌ Error during Bull Side CALL entry check: {e}")
		continue

	try:
		# --- PUT ENTRIES ---
		for stock_name, stock_data in open_high_dict.items():
			stock_id = stock_data["security_id"]
			df = sdtools.back_data(security_id=stock_id, interval=5, todays_date=back_date)
			put_df = sdtools.put_backtest_entries(df)
			put_df['stock_name'] = stock_name
			put_df['stock_id'] = stock_id
			cols = ['stock_name', 'stock_id'] + [c for c in put_df.columns if c not in ['stock_name', 'stock_id']]
			put_df = put_df[cols]
			put_df['datetime'] = pd.to_datetime(put_df['datetime']).dt.tz_localize(None)

			if not put_df.empty:
				put_df.columns = headers
				append_df_to_excel('backtest_results.xlsx', put_df, 'PUT_Entries')
			print("Put Side done")
			# print(entry_df)
	except Exception as e:
		print(f"❌ Error during Bull Side PUT entry check: {e}")
		continue
	

	# 
	

	break
	


while True:
	back_date = datetime.date(2025, 11, 14)
	reset_backtest_file(filename)
	min15_df = sdtools.get_nifty_data(interval=15)
	fib_data = sdtools.get_fib_levels(min15_df, target_date=back_date, anchor="low")
	
	print(f"📊 Fibonacci Levels for Today:")
	print(f"Range High: {fib_data['range_high']}")
	print(f"Range Low: {fib_data['range_low']}")
	print(f"Range: {fib_data['range']}")
	print(f"\n{fib_data['df']}")
	df = sdtools.get_nifty_data(interval=5)
	print(df)
	try:
		Call_entry_df = sdtools.Call_entry_check(df)
		# print(df)
		Call_entry_df['stock_name'] = "NIFTY 50"
		Call_entry_df['stock_id'] = 13
		cols = ['stock_name', 'stock_id'] + [c for c in Call_entry_df.columns if c not in ['stock_name', 'stock_id']]
		Call_entry_df = Call_entry_df[cols]
		Call_entry_df['datetime'] = pd.to_datetime(Call_entry_df['datetime']).dt.tz_localize(None)
		
		if not Call_entry_df.empty:
			Call_entry_df.columns = headers
			append_df_to_excel(filename, Call_entry_df, 'bullish_engulfing')
		print("bullish_engulfing done")

	except Exception as e:
		print(f"❌ Error during Bull Side CALL entry check: {e}")
		continue

	try:
		Put_entry_df = sdtools.Put_entry_check(df)
		# print(put_df)
		Put_entry_df['stock_name'] = "NIFTY 50"
		Put_entry_df['stock_id'] = 13
		cols = ['stock_name', 'stock_id'] + [c for c in Put_entry_df.columns if c not in ['stock_name', 'stock_id']]
		Put_entry_df = Put_entry_df[cols]
		Put_entry_df['datetime'] = pd.to_datetime(Put_entry_df['datetime']).dt.tz_localize(None)

		if not Put_entry_df.empty:
			Put_entry_df.columns = headers
			append_df_to_excel(filename, Put_entry_df, 'bearish_engulfing')
		print("bearish_engulfing done")

	except Exception as e:
		print(f"❌ Error during Bear Side PUT entry check: {e}")
		continue


	break	
'''

# backtest_stra2.py (updated bottom portion - integration with SDTools3 functions)
# -- keep the earlier top of file and helper functions as they are in your current file --
# I only show the modified/insertion part where the main loop used to run.

# ... (top of file remains unchanged; keep all existing imports and helper functions)
# Excel helper functions (existing in your file) - keep them as-is:
# reset_backtest_file, append_df_to_excel already present above in the file
# We'll ensure reset_backtest_file now creates 'NIFTY_FIB_RESULTS' sheet as well.

def reset_backtest_file(filename):
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        book = load_workbook(filename)
        for sheet in book.sheetnames:
            std = book[sheet]
            book.remove(std)

    except (FileNotFoundError, InvalidFileException, KeyError):
        book = Workbook()
        book.remove(book.active)

    headers = [
        "Entry Time", "Direction", "Fib Level", "Entry Price", "SL", "TP",
        "Exit Time", "Exit Price", "Exit Reason", "PnL"
    ]
    fib_sheet = book.create_sheet("Fib_sheet")
    fib_sheet.append(headers)
    book.save(filename)
    print(f"Reset done: {filename}")

def append_df_to_excel(filename, df, sheet_name):
	from openpyxl import load_workbook

	if not os.path.isfile(filename):
		with pd.ExcelWriter(filename, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name=sheet_name, index=False)
	else:
		book = load_workbook(filename)
		startrow = book[sheet_name].max_row if sheet_name in book.sheetnames else 0
		with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
			df.to_excel(writer, sheet_name=sheet_name, index=False, header=(startrow == 0), startrow=startrow)

current_time = datetime.datetime.now().time()
headers = ["Stock Name", "Security ID", "Timestamp", "Open", "High", "Low", "Close", "Volume", "Datetime","ema_9", "Entry"]
filename = "backtest_fib_results.xlsx"

while True:
	back_date = datetime.date(2025, 12, 19)
	reset_backtest_file(filename)
	min15_df = sdtools.get_nifty_data(interval=15, back_date=back_date)
	# print(min15_df)
	fib_data = sdtools.get_fib_levels(min15_df, target_date=back_date, anchor="low")
	nifty_5_df = sdtools.get_nifty_data(interval=5, back_date=back_date)
	try:
		Call_entry_df = sdtools.Call_entry_check(nifty_5_df, fib_data)
		# print(df)
		Call_entry_df['stock_name'] = "NIFTY 50"
		Call_entry_df['stock_id'] = 13
		cols = ['stock_name', 'stock_id'] + [c for c in Call_entry_df.columns if c not in ['stock_name', 'stock_id']]
		Call_entry_df = Call_entry_df[cols]
		Call_entry_df['datetime'] = pd.to_datetime(Call_entry_df['datetime']).dt.tz_localize(None)
		
		if not Call_entry_df.empty:
			Call_entry_df.columns = headers
			append_df_to_excel(filename, Call_entry_df, 'Fib_Call_Entries')
		print("Call backtest done")

	except Exception as e:
		print(f"❌ Error during Bull Side CALL entry check: {e}")
		continue

	try:
		Put_entry_df = sdtools.Put_entry_check(nifty_5_df, fib_data)
		# print(put_df)
		Put_entry_df['stock_name'] = "NIFTY 50"
		Put_entry_df['stock_id'] = 13
		cols = ['stock_name', 'stock_id'] + [c for c in Put_entry_df.columns if c not in ['stock_name', 'stock_id']]
		Put_entry_df = Put_entry_df[cols]
		Put_entry_df['datetime'] = pd.to_datetime(Put_entry_df['datetime']).dt.tz_localize(None)

		if not Put_entry_df.empty:
			Put_entry_df.columns = headers
			append_df_to_excel(filename, Put_entry_df, 'Fib_Put_Entries')
		print("Put backtest done")

	except Exception as e:
		print(f"❌ Error during Bear Side PUT entry check: {e}")
		continue
	break	

