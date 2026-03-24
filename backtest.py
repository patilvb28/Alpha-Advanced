# from ws_ltp import run_ltp_socket
import datetime
import time
import os
import pandas as pd
import numpy as np
import pdb
import requests
from datetime import date, timedelta
import json
from datetime import time as dt_time
from ta.trend import EMAIndicator, SMAIndicator, ADXIndicator
from ta.momentum import RSIIndicator
from openpyxl import Workbook, load_workbook

from mock_engine import MockEngine	
engine = MockEngine()
from tools import Tools
tools = Tools()
from sdtools1 import SDTools1
sdtools = SDTools1()

sdtools.get_instrument_file()

def send_telegram(message):
	try:
		import requests

		TELEGRAM_BOT_TOKEN = ""
		TELEGRAM_CHAT_ID = ""
		url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
		data = {"chat_id": TELEGRAM_CHAT_ID, "text": message}
		requests.post(url, data=data)
	except Exception as e:
		print(f"❌ Failed to send Telegram alert: {e}")

message1 = "Advanced Stock Trading Bot Started"
send_telegram(message1)
print("Advanced Stock Trading Bot Started")
tools.get_instrument_file()

main_list = ["AARTIIND", "ABB", "ABCAPITAL", "ABFRL", "ADANIENSOL", "ADANIENT", "ADANIGREEN", "ADANIPORTS", "AMBUJACEM", "ANGELONE", "APLAPOLLO", "APOLLOHOSP", "ASHOKLEY", "ASIANPAINT", "ASTRAL", "ATGL", "AUBANK", "AUROPHARMA", "AXISBANK", "BAJAJ-AUTO", "BAJAJFINSV", "BAJFINANCE", "BALKRISIND", "BANDHANBNK", "BANKBARODA", "BEL", "BHARATFORG", "BHARTIARTL", "BIOCON", "BPCL", "BRITANNIA", "BSE", "BSOFT", "CAMS", "CDSL", "CGPOWER", "CHAMBLFERT", "CHOLAFIN", "CIPLA", "COALINDIA", "COFORGE", "CONCOR", "CROMPTON", "CUMMINSIND", "DABUR", "DALBHARAT", "DIVISLAB", "DIXON", "DLF", "DMART", "DRREDDY", "EICHERMOT", "ETERNAL", "EXIDEIND", "GAIL", "GLENMARK", "GODREJCP", "GODREJPROP", "GRANULES", "GRASIM", "HAL", "HAVELLS", "HCLTECH", "HDFCAMC", "HDFCBANK", "HDFCLIFE", "HEROMOTOCO", "HINDALCO", "HINDCOPPER", "HINDPETRO", "HINDUNILVR", "HINDZINC", "ICICIBANK", "ICICIGI", "ICICIPRULI", "IGL", "INDHOTEL", "INDIGO", "INDUSINDBK", "INDUSTOWER", "INFY", "IOC", "IRCTC", "IREDA", "ITC", "JINDALSTEL", "JIOFIN", "JSWSTEEL", "JUBLFOOD", "KAYNES", "KFINTECH", "KOTAKBANK", "KPITTECH", "LAURUSLABS", "LICHSGFIN", "LICI", "LODHA", "LT", "LTF", "LTIM", "LUPIN", "M&M", "M&MFIN", "MANAPPURAM", "MARICO", "MARUTI", "MAZDOCK", "MFSL", "MGL", "MOTHERSON", "MPHASIS", "MUTHOOTFIN", "NATIONALUM", "NAUKRI", "NESTLEIND", "NTPC", "NYKAA", "OBEROIRLTY", "OFSS", "OIL", "ONGC", "PEL", "PERSISTENT", "PETRONET", "PFC", "PIDILITIND", "PIIND", "PNBHOUSING", "POLYCAB", "POWERGRID", "RECLTD", "RELIANCE", "RVNL", "SBICARD", "SBILIFE", "SBIN", "SHRIRAMFIN", "SIEMENS", "SOLARINDS", "SRF", "SUNPHARMA", "TATACHEM", "TATACOMM", "TATACONSUM", "TATAELXSI", "TATAMOTORS", "TATAPOWER", "TATATECH", "TCS", "TECHM", "TIINDIA", "TITAN", "TORNTPHARM", "TORNTPOWER", "TRENT", "TVSMOTOR", "UNITDSPR", "UNOMINDA", "UPL", "VBL", "VEDL", "VOLTAS", "WIPRO", "ZYDUSLIFE"]

FnO_list = ["ABB", "ADANIENSOL", "ADANIENT", "ADANIGREEN", "ADANIPORTS", "AMBUJACEM", "APOLLOHOSP", "ASIANPAINT", "AXISBANK", "BAJAJFINSV", "BAJFINANCE", "BHARTIARTL", "BRITANNIA", "CHOLAFIN", "CIPLA", "COALINDIA", "DABUR", "DIVISLAB", "DLF", "DMART", "DRREDDY", "EICHERMOT", "GODREJCP", "GRASIM", "HAL", "HAVELLS", "HCLTECH", "HDFCBANK", "HDFCLIFE", "HEROMOTOCO", "HINDALCO", "HINDUNILVR", "ICICIBANK", "ICICIGI", "ICICIPRULI", "INDIGO", "INDUSINDBK", "INFY", "IRCTC", "JINDALSTEL", "JSWENERGY", "JSWSTEEL", "KOTAKBANK", "LICI", "LODHA", "LT", "LTIM", "M&M", "NAUKRI", "NESTLEIND", "NTPC", "PFC", "PIDILITIND", "RECLTD", "RELIANCE", "SBILIFE", "SBIN", "SHRIRAMFIN", "SIEMENS", "SUNPHARMA", "TATACONSUM", "TATAMOTORS", "TATAPOWER", "TCS", "TECHM", "TITAN", "TORNTPHARM", "TRENT", "TVSMOTOR", "UNITDSPR", "VBL", "VEDL", "ZYDUSLIFE"]

fo_list = ["INFY", "HINDALCO","NAUKRI"]

filename = "backtest_results.xlsx"

def reset_backtest_file(filename):
	try:
		# Try to load existing workbook
		book = load_workbook(filename)
		for sheet in book.sheetnames:
			std = book[sheet]
			book.remove(std)
	except FileNotFoundError:
		# If file does not exist, create new workbook
		book = Workbook()
		# Remove the default sheet
		book.remove(book.active)

	# Add fresh empty sheets
	CALL_Entries = book.create_sheet("CALL_Entries")
	PUT_Entries = book.create_sheet("PUT_Entries")

	headers = [
		"Stock Name", "Security ID", "Timestamp", "Open", "High", "Low", "Close",
		"Volume", "Datetime", "EMA_9", "EMA_21", "SMA_9", "SMA_21",
		"RSI_9", "plus_di_9", "minus_di_9", "ADX_9", "Entry"
	]
	CALL_Entries.append(headers)
	PUT_Entries.append(headers)

	book.save(filename)
	print(f"✅ Reset done: {filename}")
	
def append_df_to_excel(filename, df, sheet_name):
	from openpyxl import load_workbook

	if not os.path.isfile(filename):
		# Create new file and write the sheet
		with pd.ExcelWriter(filename, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name=sheet_name, index=False)
	else:
		# Open existing workbook
		book = load_workbook(filename)
		# Find the starting row for appending
		startrow = book[sheet_name].max_row if sheet_name in book.sheetnames else 0

		with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
			df.to_excel(writer, sheet_name=sheet_name, index=False, header=(startrow == 0), startrow=startrow)

current_time = datetime.datetime.now().time()
headers = ["Stock Name", "Security ID", "Timestamp", "Open", "High", "Low", "Close", "Volume", "Datetime", "EMA_9", "EMA_21", "SMA_9", "SMA_21", "RSI_9", "plus_di_9", "minus_di_9", "ADX_9", "Entry"]

'''
while True:
	back_date = datetime.date(2025, 9, 3)
	open_high_dict, open_low_dict = sdtools.sd_open_high_low_dicts(stock_list=FnO_list, todays_date=back_date)
	# excel_path = "backtest_results.xlsx"  # your file name
	reset_backtest_file(filename)
	try:
		# --- CALL ENTRIES ---
		for stock_name, stock_data in open_low_dict.items():
			stock_id = stock_data["security_id"]
			df = sdtools.back_data(security_id=stock_id, interval=5, todays_date=back_date)
			call_df = sdtools.call_backtest_entries(df)
			call_df['stock_name'] = stock_name
			call_df['stock_id'] = stock_id
			cols = ['stock_name', 'stock_id'] + [c for c in call_df.columns if c not in ['stock_name', 'stock_id']]
			call_df = call_df[cols]
			call_df['datetime'] = pd.to_datetime(call_df['datetime']).dt.tz_localize(None)
			
			if not call_df.empty:
				call_df.columns = headers
				append_df_to_excel(filename, call_df, 'CALL_Entries')

			print("Call Side done")
			# print(entry_df), "ICICIGI", "ABB", "GRASIM", "TATAELXSI", "OFSS", "OFSS", "TECHM", "JUBLFOOD"
	except Exception as e:
		print(f"❌ Error during Bull Side CALL entry check: {e}")
		continue

	try:
		# --- PUT ENTRIES ---
		for stock_name, stock_data in open_high_dict.items():
			stock_id = stock_data["security_id"]
			df = sdtools.back_data(security_id=stock_id, interval=5, todays_date=back_date)
			put_df = sdtools.put_backtest_entries(df)
			put_df['stock_name'] = stock_name
			put_df['stock_id'] = stock_id
			cols = ['stock_name', 'stock_id'] + [c for c in put_df.columns if c not in ['stock_name', 'stock_id']]
			put_df = put_df[cols]
			put_df['datetime'] = pd.to_datetime(put_df['datetime']).dt.tz_localize(None)

			if not put_df.empty:
				put_df.columns = headers
				append_df_to_excel('backtest_results.xlsx', put_df, 'PUT_Entries')
			print("Put Side done")
			# print(entry_df)
	except Exception as e:
		print(f"❌ Error during Bull Side PUT entry check: {e}")
		continue
	

	# 
	

	break
	
'''

while True:
	back_date = datetime.date(2025, 10, 17)
	reset_backtest_file(filename)
	df = sdtools.get_nifty_data(interval=5)
	print(df)
	pdb.set_trace()
	try:
		call_df = sdtools.call_backtest_entries(df)
		# print(df)
		call_df['stock_name'] = "NIFTY 50"
		call_df['stock_id'] = 13
		cols = ['stock_name', 'stock_id'] + [c for c in call_df.columns if c not in ['stock_name', 'stock_id']]
		call_df = call_df[cols]
		call_df['datetime'] = pd.to_datetime(call_df['datetime']).dt.tz_localize(None)
		
		if not call_df.empty:
			call_df.columns = headers
			append_df_to_excel(filename, call_df, 'CALL_Entries')
		print("Call Side done")

	except Exception as e:
		print(f"❌ Error during Bull Side CALL entry check: {e}")
		continue

	try:
		put_df = sdtools.put_backtest_entries(df)
		# print(put_df)
		put_df['stock_name'] = "NIFTY 50"
		put_df['stock_id'] = 13
		cols = ['stock_name', 'stock_id'] + [c for c in put_df.columns if c not in ['stock_name', 'stock_id']]
		put_df = put_df[cols]
		put_df['datetime'] = pd.to_datetime(put_df['datetime']).dt.tz_localize(None)

		if not put_df.empty:
			put_df.columns = headers
			append_df_to_excel('backtest_results.xlsx', put_df, 'PUT_Entries')
		print("Put Side done")

	except Exception as e:
		print(f"❌ Error during Bull Side PUT entry check: {e}")
		continue
	break